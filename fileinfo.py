import soundfile as sf
import os
import openpyxl
import wave
import shutil

from openpyxl import Workbook

workbook = Workbook()
print(workbook.sheetnames)
workbook.save(filename="infofile.xlsx")
wb = openpyxl.load_workbook("infofile.xlsx")
sheet = wb['Sheet']
pathoffolder = 'positive'
files = os.listdir(pathoffolder)
columnvalue = 1
for k in range(1, sheet.max_row + 1):
    if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
        cell = sheet.cell(row=k, column=columnvalue)
        cell.value = "filename"

        columnvalue+=1

        break
for k in range(1, sheet.max_row + 1):
    if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
        cell = sheet.cell(row=k, column=columnvalue)
        cell.value = "subtype"
        columnvalue+=1

        break
for k in range(1, sheet.max_row + 1):
    if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
        cell = sheet.cell(row=k, column=columnvalue)
        cell.value = "format"
        columnvalue+=1

        break
for k in range(1, sheet.max_row + 1):
    if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
        cell = sheet.cell(row=k, column=columnvalue)
        cell.value = "channels"
        columnvalue+=1

        break
for k in range(1, sheet.max_row + 1):
    if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
        cell = sheet.cell(row=k, column=columnvalue)
        cell.value = "format_info"
        columnvalue+=1

        break
for k in range(1, sheet.max_row + 1):
    if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
        cell = sheet.cell(row=k, column=columnvalue)
        cell.value = "frames"
        columnvalue+=1

        break
for k in range(1, sheet.max_row + 1):
    if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
        cell = sheet.cell(row=k, column=columnvalue)
        cell.value = "samplerate"
        columnvalue+=1

        break
for k in range(1, sheet.max_row + 1):
    if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
        cell = sheet.cell(row=k, column=columnvalue)
        cell.value = "mode"
        columnvalue+=1

        break
for k in range(1, sheet.max_row + 1):
    if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
        cell = sheet.cell(row=k, column=columnvalue)
        cell.value = "subtype_info"
        columnvalue+=1

        break
for k in range(1, sheet.max_row + 1):
    if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
        cell = sheet.cell(row=k, column=columnvalue)
        cell.value = "sections"
        columnvalue+=1

        break

for i in files:
    i = pathoffolder+"/"+i
    file = i

    ob = sf.SoundFile(file)

    if ob.channels == 2:
        os.remove(i)
        continue
    selected_file = wave.open(file)
    frames = selected_file.getnframes()
    rate = selected_file.getframerate()
    duration_of_selected = frames / float(rate)

    if duration_of_selected < 3:
        shutil.move(file,"positive_dustbin")
        continue

    columnvalue = 1
    for k in range(1, sheet.max_row + 2):
        if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
            cell = sheet.cell(row=k, column=columnvalue)
            cell.value = ob.name
            columnvalue+=1

            break
    for k in range(1, sheet.max_row + 2):
        if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
            cell = sheet.cell(row=k, column=columnvalue)
            cell.value = ob.subtype
            columnvalue+=1

            break
    for k in range(1, sheet.max_row + 2):
        if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
            cell = sheet.cell(row=k, column=columnvalue)
            cell.value = ob.format
            columnvalue+=1

            break
    for k in range(1, sheet.max_row + 2):
        if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
            cell = sheet.cell(row=k, column=columnvalue)
            cell.value = ob.channels
            columnvalue+=1

            break
    for k in range(1, sheet.max_row + 2):
        if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
            cell = sheet.cell(row=k, column=columnvalue)
            cell.value = ob.format_info
            columnvalue+=1

            break
    for k in range(1, sheet.max_row +2 ):
        if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
            cell = sheet.cell(row=k, column=columnvalue)
            cell.value = ob.frames
            columnvalue+=1

            break
    for k in range(1, sheet.max_row + 2):
        if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
            cell = sheet.cell(row=k, column=columnvalue)
            cell.value = ob.samplerate
            columnvalue+=1

            break
    for k in range(1, sheet.max_row + 2):
        if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
            cell = sheet.cell(row=k, column=columnvalue)
            cell.value = ob.mode
            columnvalue+=1

            break
    for k in range(1, sheet.max_row + 2):
        if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
            cell = sheet.cell(row=k, column=columnvalue)
            cell.value = ob.subtype_info
            columnvalue+=1

            break
    for k in range(1, sheet.max_row + 2):
        if str(sheet.cell(row=k, column=columnvalue).value) == 'None':
            cell = sheet.cell(row=k, column=columnvalue)
            cell.value = ob.sections
            columnvalue+=1

            break

wb.save("infofile.xlsx")









# Overall, this script is designed to analyze sound files in a specified folder, extract their metadata,--
# and store that metadata in an Excel file for further analysis or processing.

# openpyxl -- it is use to create an excel file


# The os library in Python provides a way to interact with the operating system, allowing you to perform various tasks --
# such as file and directory operations, process management, and environment variables manipulation



# The wave library in Python provides functionality for reading and writing WAV (Waveform Audio File Format) files,--
# which are commonly used for storing uncompressed audio data.


# The shutil library in Python provides a high-level interface for file operations, allowing you to perform various--
# file-related tasks such as copying, moving, and deleting files and directories. 




# Overall, the soundfile library is an essential tool for working with sound files in Python, providing a comprehensive and--
# efficient solution for reading, writing, and manipulating audio data across various formats.










# --fn of the code 


# This Python script, named "fileinfo.py," performs several tasks related to processing sound files and storing their metadata in an Excel file:

# 1. It imports necessary libraries such as `soundfile`, `os`, `openpyxl`, `wave`, and `shutil`.

# 2. It creates a new Excel workbook named "infofile.xlsx" using the `openpyxl` library.

# 3. It loads the existing workbook named "infofile.xlsx" using `openpyxl`.

# 4. It defines the path to the folder containing sound files (`positive`).

# 5. It iterates through the files in the specified folder.

# 6. For each file, it checks if the number of audio channels is 2 (stereo). If so, it removes the file from the folder.

# 7. It calculates the duration of the audio file and moves files with a duration of less than 3 seconds to a separate folder named "positive_dustbin."

# 8. It iterates through the metadata attributes of each sound file (e.g., filename, subtype, format, channels, format_info, frames, samplerate, mode, subtype_info, sections).

# 9. For each attribute, it searches for the first empty cell in the corresponding column of the Excel sheet and writes the metadata value to that cell.

# 10. After processing all sound files, it saves the updated Excel workbook.

# Overall, this script is designed to analyze sound files in a specified folder, extract their metadata, and store that metadata in an Excel file for further analysis or processing.